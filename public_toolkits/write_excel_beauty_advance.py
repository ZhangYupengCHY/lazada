#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/8/6 0006 14:43
# @Author  : Zhang YP
# @Email   : 1579922399@qq.com
# @github  :  Aaron Ramsey
# @File    : write_excel_beauty.py

"""
    将df好看的写入到xlsx中
"""
import os

import pandas as pd
import openpyxl  # Connect the library
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, NamedStyle  # Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill  # Connect styles for text
from openpyxl.styles import colors
from openpyxl.utils import get_column_letter


class ExcelWriterBeauty(object):
    """
        调整df输出到单元格格式:
            单元格大小(宽,高)
            单元格背景颜色
            单元格字体大小,字体,字体颜色,是否是斜体,是否有下划线,是否加粗
    """

    def __init__(self, save_path):
        """
            初始化保存的文件路径
        :param save_path: dir
            文件保存路径
        """
        if not os.path.exists(os.path.dirname(save_path)):
            raise FileNotFoundError(f'{save_path} is not exist.')
        self._save_path = save_path
        self._book = openpyxl.Workbook()

    def write_excel(self, df: pd.DataFrame, sheet_name='Sheet1', min_cell_width=12, max_cell_width=100,
                    head_height=3, content_height=2,
                    summary_row=False,
                    head_options={'font_color': [255, 255, 255], 'font-family': 'Microsoft YaHei',
                                  'font-size': 13, 'font-bord': True, 'font-underline': 'none',
                                  'font-italic': False, 'background-color': [0, 51, 102]},
                    content_options={'font_color': [0, 0, 0], 'font-family': 'Microsoft YaHei',
                                     'font-size': 13, 'font-bord': False, 'font-underline': 'none',
                                     'font-italic': False, 'background-color-odd': [245, 248, 252],
                                     'background-color-even': [225, 232, 240]}):
        """
            将df按照格式路径xlsx的工作表中
            将工作表格式分为三个部分:标题行、数据行、汇总行(如果存在)
            格式项:
                    单元格背景颜色,单元格边框,单元格高度,单元格宽度,字体,字体颜色,字体加粗,字体大小,字体上下对齐方式,字体左右对齐方式
            待开发的格式项:
                边框:边框大小,颜色,样式等
        :parma head_options: dict
            工作表表头选项:
                font_color:list [r,g,b]
                    字体颜色
                font-family: str
                    字体
                font-size:int
                    字体大小
                font-bord：bool
                    字体是否加粗
                font-underline:bool
                    字体是否拥有下划线
                font-italic:bool
                    是否是斜体
                background-color: list [r,g,b]
                    背景颜色
        :parma content_option:dict
            same as head_options params
        :param summary_row: bool default False
            数据结尾是否拥有总结行
        :param head_height: int default 3
            表头单元格高度
        :param content_height: int default 2
            内容单元格高度
        :param min_cell_width: int defalut 12
            最小单元格宽度
        :param max_cell_width: int defalut 100
            最大单元格宽度
        :param df: pd.DataFrame
        :param sheet_name:str
            保存的工作表名

        :param save_path:write into path
            保存写入的路径
        :return: None
        """

        # 对df进行检测
        if df is None:
            raise TypeError(f'import is None.')
        if not isinstance(df, pd.DataFrame):
            raise TypeError(f'import type must be pd.DataFrame not {type(df)}.')
        if df.empty:
            return

        # 对输入参数进行检测:
        for int_value in [min_cell_width, max_cell_width, head_height, content_height]:
            if not isinstance(int_value, int):
                raise TypeError(f'cell width or height must be int.import {int_value} is not int.')

        for dict_value in [head_options, content_options]:
            if not isinstance(dict_value, dict):
                raise TypeError(f'cell options must be dict.import {dict_value} is not dict.')

        # 检测rgb格式
        for rgb_list in [head_options['font_color'], head_options['background-color'], content_options['font_color'],
                         content_options['background-color-odd'], content_options['background-color-even']]:
            if not isinstance(rgb_list, list):
                raise TypeError(f'Cell font-color or background-color must be RGB list.import {rgb_list} not list.')
            if len(rgb_list) != 3:
                raise ValueError(f'RGB must element is 3,but you rgb is {len(rgb_list)}.')
            for value in rgb_list:
                if value < 0 or value > 255:
                    raise ValueError(f'RGB must 0~255,but you rgb value is {value}.')

        # 初始化工作表
        if 'Sheet' in self._book.sheetnames:
            sheet = self._book.active
            sheet.title = sheet_name
        else:
            sheet = self._book.create_sheet()
            sheet.title = sheet_name

        # rgb数组转换成十六进制格式
        def rgb_list_2_hex(rgb_list):
            if not isinstance(rgb_list, list):
                raise TypeError(f'{rgb_list} not list.')
            if len(rgb_list) != 3:
                raise ValueError(f'RGB must element is 3,but you rgb is {len(rgb_list)}.')
            for value in rgb_list:
                if value < 0 or value > 255:
                    raise ValueError(f'RGB must 0~255,but you rgb value is {value}.')
            hex_color = ''
            for item in rgb_list:
                item_hex_color = hex(item)[2:].upper()
                if len(item_hex_color) == 1:
                    item_hex_color = f'0{item_hex_color}'
                hex_color += item_hex_color
            return hex_color

        # 设置字体

        head_font = Font(size=head_options['font-size'], color=f"{rgb_list_2_hex(head_options['font_color'])}",
                         name=head_options['font-family'],
                         bold=head_options['font-bord'], italic=head_options['font-italic'],
                         underline=head_options['font-underline'])

        content_font = Font(size=content_options['font-size'], color=f"{rgb_list_2_hex(content_options['font_color'])}",
                            name=content_options['font-family'],
                            bold=content_options['font-bord'], italic=content_options['font-italic'],
                            underline=content_options['font-underline'])

        # 设置背景颜色
        head_background_color = PatternFill(start_color=f"{rgb_list_2_hex(head_options['background-color'])}",
                                            fill_type="solid")
        content_background_color_odd = PatternFill(
            start_color=f"{rgb_list_2_hex(content_options['background-color-odd'])}", fill_type="solid")
        content_background_color_even = PatternFill(
            start_color=f"{rgb_list_2_hex(content_options['background-color-even'])}", fill_type="solid")

        # 设置对齐方式
        head_alignment = Alignment(horizontal='left', vertical='center')
        content_alignment = Alignment(horizontal='center', vertical='center')

        #  设置样式
        # 标题样式
        head_style = NamedStyle(name='head')
        head_style.font = head_font
        head_style.alignment = head_alignment
        head_style.fill = head_background_color
        # 设置数据表内容样式
        # 奇数项
        content_style_odd = NamedStyle(name='content_odd')
        content_style_odd.font = content_font
        content_style_odd.alignment = content_alignment
        content_style_odd.fill = content_background_color_odd

        # 奇数项
        content_style_even = NamedStyle(name='content_even')
        content_style_even.font = content_font
        content_style_even.alignment = content_alignment
        content_style_even.fill = content_background_color_even

        # 注册样式表
        new_styles_dict = {'head': head_style, 'content_odd': content_style_odd, 'content_even': content_style_even}
        for style_name, style_value in new_styles_dict.items():
            if style_name not in self._book.named_styles:
                self._book.add_named_style(style_value)

        # 5.设置列宽，一个中文等于两个英文等于两个字符，11为字符数，256为衡量单位
        def get_col_width(col_list, extent_len=1.4, min_cell_width=min_cell_width,
                          max_cell_width=max_cell_width) -> int:
            """
                excel中文占用两个字符,英文和数字占用一个字符
                判断最大值是否为异常值:
                    如果不是异常值,则用最大值excel的宽度的1.2倍为列宽
                    如果是异常值,则用Q3+1.5IQR的1.2倍作为excel的列宽

            :param col_list:
            :param extent_len:int
                 自适应扩展宽度倍数
            :return: int
                 计算出写入到excel中的宽度
            """

            # 计算字符串长度
            def calc_str_width(calc_str):
                # 计算字符串的长度:中文等为2个字符,数字和英文为1个字符
                if not isinstance(calc_str, str):
                    if pd.isna(calc_str):
                        calc_str = ' '
                    else:
                        calc_str = str(calc_str)
                ori_str_len = len(calc_str)
                chinses_num = 0
                for i in calc_str:
                    # 中文检测:中文符号区
                    if '\u4e00' <= i <= '\u9fff':
                        chinses_num += 1
                str_len = ori_str_len + chinses_num
                return str_len

            if len(col_list) == 0:
                return 1
            # 计算List中每个占用字符长度
            list_element_len = [calc_str_width(str) for str in col_list]
            list_element_len_sorted = sorted(list_element_len)
            list_len = len(col_list)
            # 计算list的最大长度和上四分位数以及四分位间距
            list_element_max_len = list_element_len_sorted[-1]
            # 上四分位数
            q3_element_max_len = list_element_len_sorted[int(0.75 * list_len)]
            # 下四分位数
            q1_element_max_len = list_element_len_sorted[int(0.25 * list_len)]
            # 1.5倍四分位间距
            # 有效上线
            limit_max_len = q3_element_max_len + 1.5 * (q3_element_max_len - q1_element_max_len)
            cell_len = min(limit_max_len, list_element_max_len)
            return min(max_cell_width, max(int(extent_len * cell_len), min_cell_width))

        i = 1
        for _, col_value in df.iteritems():
            col_len = get_col_width(col_value, min_cell_width=min_cell_width, max_cell_width=max_cell_width)
            column_name = get_column_letter(i)
            sheet.column_dimensions[column_name].width = col_len
            i += 1

        # 6.设置行高
        # 首行行高
        # first you should tell xlwt row height and default font height do not match
        sheet.row_dimensions[1].height = head_height * 20
        # 设置其他行
        for i in range(2, len(df) + 2):
            sheet.row_dimensions[i].height = content_height * 20
        if summary_row:
            sheet.row_dimensions[len(df)+1].height = head_height * 20

        def row_writer(row_value, row_index, sheet, style):
            for i, value in enumerate(row_value):
                sheet.cell(row=row_index + 1, column=i + 1).value = value
                sheet.cell(row=row_index + 1, column=i + 1).style = style

        # 写列标题
        column = df.columns
        columns_index = 0
        row_writer(column, columns_index, sheet, 'head')
        # 写内容
        df.reset_index(drop=True, inplace=True)
        for row, row_value in df.iterrows():
            if (row % 2) == 0:
                row_writer(row_value, row + 1, sheet, 'content_odd')
            else:
                row_writer(row_value, row + 1, sheet, 'content_even')

        if summary_row:
            row_writer(list(df.tail(1).values[0]), len(df), sheet, 'head')

    def save(self):
        self._book.save(self._save_path)


if __name__ == '__main__':
    data = pd.read_csv(data_path)
    a = ExcelWriterBeauty(save_path)
    a.write_excel(data, sheet_name='so', summary_row=False)
    a.write_excel(data, sheet_name='what', summary_row=False)
    a.save()
